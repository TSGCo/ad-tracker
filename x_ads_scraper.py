import re
import requests
import pandas as pd
import zipfile
import io
import traceback
from datetime import datetime, timedelta
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

X_CSV_ROW_CAP = 80_000
X_CSV_CHUNK_SIZE = 15_000
X_STREAMING_MATCH_CAP = 50_000

X_DATA_BASE_URL = "https://business.x.com/content/dam/business-twitter/political-ads-data"
X_POLITICAL_ADS_DISCLOSURE_URL = (
    "https://business.x.com/en/help/ads-policies/ads-content-policies/"
    "political-content/political-ads-disclosure"
)

_REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
}

_RE_DISCLOSURE_ZIP = re.compile(
    r'/content/dam/business-twitter/political-ads-data/'
    r'(\d{1,2}-[A-Za-z]+-\d{4}-political-ads-data\.zip)',
    re.IGNORECASE,
)

_MONTH_TO_NUM = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}

_UNNAMED_COLUMN_PATTERN = re.compile(r"^Unnamed:\s*\d+$", re.IGNORECASE)

STATE_MAPPING = {
    'al': 'alabama', 'ak': 'alaska', 'az': 'arizona', 'ar': 'arkansas',
    'ca': 'california', 'co': 'colorado', 'ct': 'connecticut', 'de': 'delaware',
    'fl': 'florida', 'ga': 'georgia', 'hi': 'hawaii', 'id': 'idaho',
    'il': 'illinois', 'in': 'indiana', 'ia': 'iowa', 'ks': 'kansas',
    'ky': 'kentucky', 'la': 'louisiana', 'me': 'maine', 'md': 'maryland',
    'ma': 'massachusetts', 'mi': 'michigan', 'mn': 'minnesota', 'ms': 'mississippi',
    'mo': 'missouri', 'mt': 'montana', 'ne': 'nebraska', 'nv': 'nevada',
    'nh': 'new hampshire', 'nj': 'new jersey', 'nm': 'new mexico', 'ny': 'new york',
    'nc': 'north carolina', 'nd': 'north dakota', 'oh': 'ohio', 'ok': 'oklahoma',
    'or': 'oregon', 'pa': 'pennsylvania', 'ri': 'rhode island', 'sc': 'south carolina',
    'sd': 'south dakota', 'tn': 'tennessee', 'tx': 'texas', 'ut': 'utah',
    'vt': 'vermont', 'va': 'virginia', 'wa': 'washington', 'wv': 'west virginia',
    'wi': 'wisconsin', 'wy': 'wyoming', 'dc': 'district of columbia'
}


def _filename_sort_key(zip_filename: str) -> tuple[int, int, int]:
    m = re.match(
        r"^(\d{1,2})-([A-Za-z]+)-(\d{4})-political-ads-data\.zip$",
        zip_filename,
        re.IGNORECASE,
    )
    if not m:
        return (0, 0, 0)
    day_s, month_s, year_s = m.group(1), m.group(2).lower(), m.group(3)
    month_n = _MONTH_TO_NUM.get(month_s)
    if month_n is None:
        return (0, 0, 0)
    return (int(year_s), month_n, int(day_s))


def find_latest_data_file_from_disclosure_page():
    try:
        r = requests.get(
            X_POLITICAL_ADS_DISCLOSURE_URL,
            timeout=20,
            headers=_REQUEST_HEADERS,
            allow_redirects=True,
        )
        if r.status_code != 200:
            logger.warning("Disclosure page HTTP %s", r.status_code)
            return None, None
        names = _RE_DISCLOSURE_ZIP.findall(r.text)
        if not names:
            logger.warning("No political-ads-data zip links found on disclosure page")
            return None, None
        best_name = max(names, key=_filename_sort_key)
        url = f"https://business.x.com/content/dam/business-twitter/political-ads-data/{best_name}"
        label = best_name.replace("-political-ads-data.zip", "")
        logger.info("Using zip from disclosure page: %s", best_name)
        return url, label
    except requests.RequestException as e:
        logger.warning("Could not fetch disclosure page: %s", e)
        return None, None


def generate_possible_dates(days_back=30):
    dates = []
    today = datetime.now()

    seen = set()
    for i in range(days_back):
        date = today - timedelta(days=i)
        month_lower = date.strftime("%B").lower()
        for day_fmt in (f"{date.day:02d}", f"{date.day}"):
            formatted_date = f"{day_fmt}-{month_lower}-{date.year}"
            if formatted_date in seen:
                continue
            seen.add(formatted_date)
            dates.append((formatted_date, date))

    return dates


def find_latest_data_file():
    url, label = find_latest_data_file_from_disclosure_page()
    # Trust the official disclosure page link. (HEAD often returns 404 while GET returns 200.)
    if url:
        return url, label

    possible_dates = generate_possible_dates(days_back=30)

    for date_str, date_obj in possible_dates:
        try_url = f"{X_DATA_BASE_URL}/{date_str}-political-ads-data.zip"

        try:
            logger.info(f"Checking for file: {date_str}")
            response = requests.get(
                try_url,
                timeout=10,
                stream=True,
                allow_redirects=True,
                headers=_REQUEST_HEADERS,
            )
            response.close()

            if response.status_code == 200:
                logger.info(f"Found latest data file: {date_str}")
                return try_url, date_str
        except requests.RequestException as e:
            logger.debug(f"Date {date_str} not found: {e}")
            continue

    logger.warning("Could not find any recent X political ads data file")
    return None, None


def _read_xlsx_streaming(xlsx_bytes, has_search, advertiser_name, geography):
    from openpyxl import load_workbook

    bio = io.BytesIO(xlsx_bytes)
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            logger.warning("XLSX sheet is empty")
            return pd.DataFrame()

        header = [
            str(h).strip() if h is not None and str(h).strip() else f"Unnamed: {i}"
            for i, h in enumerate(header_row)
        ]
        seen = {}
        uniq_header = []
        for h in header:
            if h in seen:
                seen[h] += 1
                uniq_header.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 0
                uniq_header.append(h)
        header = uniq_header
        n_cols = len(header)

        def _normalize_row(row):
            lst = list(row) if row else []
            if len(lst) < n_cols:
                lst = list(lst) + [None] * (n_cols - len(lst))
            elif len(lst) > n_cols:
                lst = lst[:n_cols]
            return lst

        buffer = []
        total_read = 0
        matches = []
        browse_parts = []

        def _flush_buffer():
            nonlocal buffer, total_read, matches, browse_parts
            if not buffer:
                return
            if has_search and sum(len(m) for m in matches) >= X_STREAMING_MATCH_CAP:
                buffer = []
                return
            if not has_search and sum(len(p) for p in browse_parts) >= X_CSV_ROW_CAP:
                buffer = []
                return
            chunk = pd.DataFrame([_normalize_row(r) for r in buffer], columns=header)
            buffer = []
            total_read += len(chunk)

            if has_search:
                chunk = standardize_columns(chunk)
                chunk = filter_by_advertiser(chunk, advertiser_name or "")
                if geography and "Geography Targeting" in chunk.columns:
                    try:
                        expanded_geo = expand_geography_search(geography)
                        mask = chunk["Geography Targeting"].astype(str).str.contains(
                            expanded_geo, case=False, na=False, regex=True
                        )
                        chunk = chunk[mask]
                    except Exception:
                        chunk = chunk[
                            chunk["Geography Targeting"].astype(str).str.lower().str.contains(
                                geography.lower(), na=False
                            )
                        ]
                if not chunk.empty:
                    matches.append(chunk)
            else:
                browse_parts.append(chunk)

        for row in rows_iter:
            buffer.append(row)
            if len(buffer) >= X_CSV_CHUNK_SIZE:
                _flush_buffer()
                if has_search and sum(len(m) for m in matches) >= X_STREAMING_MATCH_CAP:
                    break
                if not has_search and sum(len(p) for p in browse_parts) >= X_CSV_ROW_CAP:
                    break

        _flush_buffer()

        if has_search:
            df = (
                pd.concat(matches, ignore_index=True).head(X_STREAMING_MATCH_CAP)
                if matches
                else pd.DataFrame()
            )
            logger.info("[5/7] XLSX streaming done: rows_scanned=%s, matches=%s", total_read, len(df))
        else:
            if not browse_parts:
                df = pd.DataFrame()
            else:
                df = pd.concat(browse_parts, ignore_index=True).head(X_CSV_ROW_CAP)
                df = standardize_columns(df)
            logger.info("[5/7] XLSX read done, shape=%s", df.shape)

        return df
    finally:
        wb.close()


def download_and_extract_csv(advertiser_name=None, geography=None):
    has_search = bool(advertiser_name or geography)
    logger.info("[1/7] find_latest_data_file()")
    url, date_str = find_latest_data_file()

    if not url:
        raise Exception("Could not find latest X political ads data file")
    logger.info("[2/7] Downloading response")
    try:
        logger.info(f"Downloading X political ads data from: {url}")
        response = requests.get(url, timeout=30, headers=_REQUEST_HEADERS)
        response.raise_for_status()
        logger.info(f"[2/7] Response OK, content length={len(response.content)} bytes")

        logger.info("[3/7] Extracting file from ZIP")
        with zipfile.ZipFile(io.BytesIO(response.content)) as zip_file:
            file_list = zip_file.namelist()
            logger.info(f"Files in ZIP: {file_list}")

            csv_files = [f for f in file_list if f.endswith('.csv') and not f.startswith('__MACOSX')]
            xlsx_files = [f for f in file_list if f.endswith('.xlsx') and not f.startswith('__MACOSX')]
            dataframe_from_csv_browse = False

            if csv_files:
                file_path = csv_files[0]
                with zip_file.open(file_path) as zf:
                    if has_search:
                        logger.info(f"[4/7] Streaming CSV with keyword=%r, geography=%r (full-file search)", advertiser_name, geography)
                        matches = []
                        total_rows_read = 0
                        for i, chunk in enumerate(
                            pd.read_csv(
                                zf,
                                encoding="utf-8",
                                on_bad_lines="skip",
                                low_memory=False,
                                chunksize=X_CSV_CHUNK_SIZE,
                            )
                        ):
                            total_rows_read += len(chunk)
                            if (i + 1) % 20 == 0:
                                logger.info("[5/7] Streaming chunk %s, rows read=%s, matches so far=%s", i + 1, total_rows_read, sum(len(m) for m in matches))
                            chunk = standardize_columns(chunk)
                            chunk = filter_by_advertiser(chunk, advertiser_name or "")
                            if geography and "Geography Targeting" in chunk.columns:
                                try:
                                    expanded_geo = expand_geography_search(geography)
                                    mask = chunk["Geography Targeting"].astype(str).str.contains(
                                        expanded_geo, case=False, na=False, regex=True
                                    )
                                    chunk = chunk[mask]
                                except Exception:
                                    chunk = chunk[
                                        chunk["Geography Targeting"].astype(str).str.lower().str.contains(
                                            geography.lower(), na=False
                                        )
                                    ]
                            if not chunk.empty:
                                matches.append(chunk)
                            if sum(len(m) for m in matches) >= X_STREAMING_MATCH_CAP:
                                break
                        df = pd.concat(matches, ignore_index=True).head(X_STREAMING_MATCH_CAP) if matches else pd.DataFrame()
                        logger.info(f"[5/7] Streaming done: rows_scanned=%s, matches=%s", total_rows_read, len(df))
                    else:
                        logger.info(f"[4/7] Opening ZIP entry: {file_path} (capped at {X_CSV_ROW_CAP} rows)")
                        df = pd.read_csv(
                            zf,
                            encoding="utf-8",
                            on_bad_lines="skip",
                            low_memory=False,
                            nrows=X_CSV_ROW_CAP,
                        )
                        logger.info(f"[5/7] pd.read_csv() done, shape={df.shape}")
                        dataframe_from_csv_browse = True

            elif xlsx_files:
                file_path = xlsx_files[0]
                logger.info(f"[4/7] Reading XLSX (streaming): {file_path}")
                with zip_file.open(file_path) as f:
                    xlsx_bytes = f.read()
                df = _read_xlsx_streaming(xlsx_bytes, has_search, advertiser_name, geography)

            else:
                raise Exception(f"No CSV or XLSX files found in ZIP. Contents: {file_list}")

        if dataframe_from_csv_browse and not df.empty:
            df = standardize_columns(df)
        logger.info(f"[6/7] Successfully loaded {len(df)} rows from X political ads data")
        logger.info("[7/7] Returning dataframe from download_and_extract_csv")
        return df

    except requests.RequestException as e:
        logger.error(f"Error downloading file: {e}")
        raise Exception(f"Failed to download X political ads data: {e}") from e
    except zipfile.BadZipFile as e:
        logger.error(f"Error extracting ZIP: {e}")
        raise Exception(f"Failed to extract ZIP file: {e}") from e
    except Exception as e:
        logger.exception("X ads CSV load failed: %s", e)
        traceback.print_exc()
        raise


def filter_by_advertiser(df, keyword):
    logger.info("filter_by_advertiser: keyword=%r, input_rows=%s", keyword, len(df) if df is not None else None)
    if not keyword:
        return df

    needle = (keyword or "").strip()
    if not needle:
        return df

    search_columns = [col for col in df.columns if col.lower() in [
        'advertiser name', 'screen name', 'ad type', 'ad id', 'ad url'
    ]]

    if not search_columns:
        logger.warning("Could not find searchable columns. Available columns: " + str(df.columns.tolist()))
        return df

    mask = False
    for col in search_columns:
        if col in df.columns:
            mask = mask | df[col].astype(str).str.contains(
                needle, case=False, na=False, regex=False
            )

    filtered_df = df[mask]
    logger.info("filter_by_advertiser: output_rows=%s", len(filtered_df))
    return filtered_df


def _coerce_ad_id_for_arrow(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty or "Ad Id" not in df.columns:
        return df

    def _cell(x):
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return pd.NA
        if isinstance(x, str):
            s = x.strip()
            return s if s else pd.NA
        if isinstance(x, bool):
            return str(x)
        if isinstance(x, int):
            return str(x)
        if isinstance(x, float):
            return str(int(x)) if x.is_integer() else str(x)
        return str(x)

    out = df.copy()
    out["Ad Id"] = out["Ad Id"].map(_cell).astype("string")
    return out


def drop_unnamed_junk_columns(df):
    if df is None or df.empty:
        return df
    to_drop = [
        c
        for c in df.columns
        if isinstance(c, str) and _UNNAMED_COLUMN_PATTERN.match(c.strip())
    ]
    if to_drop:
        df = df.drop(columns=to_drop, errors="ignore")
        logger.info("drop_unnamed_junk_columns: removed %s columns", len(to_drop))
    return df


def expand_geography_search(geography_query):
    if not geography_query:
        return geography_query
    
    query_lower = geography_query.lower().strip()
    
    if query_lower in STATE_MAPPING:
        full_name = STATE_MAPPING[query_lower]
        return f"(?:{query_lower}|{full_name})"

    for abbr, full_name in STATE_MAPPING.items():
        if query_lower == full_name:
            return f"(?:{abbr}|{full_name})"
    
    return geography_query


def standardize_columns(df):
    cols = list(df.columns) if df is not None else []
    cols_preview = cols[:15] if len(cols) > 15 else cols
    if len(cols) > 15:
        cols_preview.append("...")
    logger.info("standardize_columns: input shape=%s, columns=%s", df.shape if df is not None else None, cols_preview)
    column_mapping = {
        'Screen Name': 'Advertiser Name',
        'Tweet Id': 'Ad Id',
        'Tweet Url': 'Ad Url',
        'Day of Start Date Adgroup': 'Start Date',
        'Day of End Date Adgroup': 'End Date',
        'Targeting Name': 'Ad Type',
        'Interest Targeting': 'Interest Targeting',
        'Geo Targeting': 'Geography Targeting',
        'Gender Targeting': 'Gender Targeting',
        'Age Targeting': 'Age Targeting',
        'Impressions': 'Impressions',
        'Spend_USD': 'Spend',
    }
    
    rename_dict = {}
    for old_col, new_col in column_mapping.items():
        if old_col in df.columns:
            rename_dict[old_col] = new_col
    
    df = df.rename(columns=rename_dict)
    df = drop_unnamed_junk_columns(df)
    df = _coerce_ad_id_for_arrow(df)
    out_cols = list(df.columns)
    out_preview = out_cols[:15] if len(out_cols) > 15 else out_cols
    if len(out_cols) > 15:
        out_preview.append("...")
    logger.info("standardize_columns: done, output columns=%s", out_preview)
    return df
